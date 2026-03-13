"""
文件管理 API 路由。

端点:
- POST /api/files/upload        — 上传文件
- GET  /api/files/              — 列出用户文件
- GET  /api/files/{file_id}     — 文件元数据
- GET  /api/files/{file_id}/text — 提取文本
- GET  /api/files/{file_id}/download — 下载文件
- GET  /api/files/{file_id}/preview  — 结构化预览数据
- GET  /api/files/{file_id}/render   — 原样返回 HTML (iframe)
- DELETE /api/files/{file_id}   — 删除文件
"""

from __future__ import annotations

import logging
from pathlib import Path
from urllib.parse import quote
from typing import Optional
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import JSONResponse, Response

from core.auth import AuthUser, get_current_user
from dependencies import get_file_service

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/files", tags=["files"])


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    session_id: Optional[str] = Query(None),
    user: AuthUser = Depends(get_current_user),
):
    """上传文件到用户空间。"""
    service = get_file_service()

    try:
        content = await file.read()
        filename = file.filename or "unnamed"

        metadata = service.save_file(
            user.tenant_id, user.user_id, filename, content,
            session_id=session_id or "",
        )

        return {
            "file_id": metadata.file_id,
            "filename": metadata.filename,
            "content_type": metadata.content_type,
            "size_bytes": metadata.size_bytes,
            "sha256": metadata.sha256,
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"Upload failed: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {e}")


@router.get("/")
async def list_user_files(
    session_id: Optional[str] = Query(None),
    user: AuthUser = Depends(get_current_user),
):
    """列出用户所有文件，可按 session_id 过滤。"""
    service = get_file_service()

    if session_id:
        files = service.list_files_by_session(user.tenant_id, user.user_id, session_id)
    else:
        files = service.list_files(user.tenant_id, user.user_id)
    return {
        "user_id": user.user_id,
        "files": [
            {
                "file_id": f.file_id,
                "filename": f.filename,
                "content_type": f.content_type,
                "size_bytes": f.size_bytes,
                "created_at": f.created_at,
                "session_id": f.session_id,
            }
            for f in files
        ],
    }


@router.get("/{file_id}")
async def get_file_metadata(file_id: str, user: AuthUser = Depends(get_current_user)):
    """获取文件元数据。"""
    service = get_file_service()

    try:
        metadata, _ = service.get_file(user.tenant_id, user.user_id, file_id)
        return {
            "file_id": metadata.file_id,
            "filename": metadata.filename,
            "content_type": metadata.content_type,
            "size_bytes": metadata.size_bytes,
            "sha256": metadata.sha256,
            "created_at": metadata.created_at,
        }
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"File not found: {file_id}")


@router.get("/{file_id}/text")
async def get_extracted_text(file_id: str, user: AuthUser = Depends(get_current_user)):
    """提取文件文本内容。"""
    service = get_file_service()

    try:
        text = service.extract_text(user.tenant_id, user.user_id, file_id)
        return {
            "file_id": file_id,
            "text": text,
        }
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"File not found: {file_id}")


@router.get("/{file_id}/download")
async def download_file(file_id: str, user: AuthUser = Depends(get_current_user)):
    """下载文件（返回原始文件内容）。"""
    service = get_file_service()

    try:
        metadata, content = service.get_file(user.tenant_id, user.user_id, file_id)
        # RFC 5987: 中文文件名用 UTF-8 编码
        encoded_name = quote(metadata.filename)
        return Response(
            content=content,
            media_type=metadata.content_type or "application/octet-stream",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
                "Content-Length": str(len(content)),
            },
        )
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"File not found: {file_id}")


@router.get("/{file_id}/preview")
async def preview_file(file_id: str, user: AuthUser = Depends(get_current_user)):
    """返回结构化预览数据。"""
    service = get_file_service()

    try:
        metadata, content = service.get_file(user.tenant_id, user.user_id, file_id)
    except FileNotFoundError:
        return JSONResponse(status_code=404, content={"error": "File not found"})

    ext = Path(metadata.filename).suffix.lower()
    base_url = f"/api/files/{file_id}"

    # Image
    if ext in ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.svg'):
        return {"type": "image", "url": f"{base_url}/download", "filename": metadata.filename}

    # PDF
    if ext == '.pdf':
        return {"type": "pdf", "url": f"{base_url}/download", "filename": metadata.filename}

    # DOCX
    if ext in ('.docx', '.doc'):
        text = service.extract_text(user.tenant_id, user.user_id, file_id)
        return {"type": "docx", "content": text, "filename": metadata.filename}

    # Excel
    if ext in ('.xlsx', '.xls'):
        try:
            sheets = _extract_excel(content)
            return {"type": "excel", "sheets": sheets, "filename": metadata.filename}
        except Exception as e:
            return {"type": "text", "content": f"[Excel 解析失败: {e}]", "filename": metadata.filename}

    # HTML
    if ext == '.html':
        try:
            source = content.decode('utf-8')
        except UnicodeDecodeError:
            source = content.decode('utf-8', errors='replace')
        return {"type": "html", "source": source, "render_url": f"{base_url}/render", "filename": metadata.filename}

    # Code files
    code_exts = {
        '.py': 'python', '.js': 'javascript', '.ts': 'typescript',
        '.jsx': 'jsx', '.tsx': 'tsx', '.css': 'css', '.scss': 'scss',
        '.java': 'java', '.go': 'go', '.rs': 'rust', '.c': 'c',
        '.cpp': 'cpp', '.h': 'c', '.rb': 'ruby', '.php': 'php',
        '.sh': 'bash', '.sql': 'sql', '.yaml': 'yaml', '.yml': 'yaml',
        '.xml': 'xml', '.svg': 'xml',
    }
    if ext in code_exts:
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            text = content.decode('utf-8', errors='replace')
        return {"type": "code", "content": text, "language": code_exts[ext], "filename": metadata.filename}

    # Markdown
    if ext == '.md':
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            text = content.decode('utf-8', errors='replace')
        return {"type": "markdown", "content": text, "filename": metadata.filename}

    # Plain text
    if ext in ('.txt', '.csv', '.json', '.log', '.ini', '.conf'):
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                text = content.decode('gbk')
            except UnicodeDecodeError:
                text = content.decode('utf-8', errors='replace')
        lang = 'json' if ext == '.json' else 'csv' if ext == '.csv' else ''
        return {"type": "code" if lang else "text", "content": text, "language": lang, "filename": metadata.filename}

    # Fallback
    return {
        "type": "unsupported",
        "filename": metadata.filename,
        "content_type": metadata.content_type,
        "size_bytes": metadata.size_bytes,
    }


@router.get("/{file_id}/render")
async def render_file(file_id: str, user: AuthUser = Depends(get_current_user)):
    """原样返回 HTML 文件 (用于 iframe 加载)。"""
    service = get_file_service()
    try:
        metadata, content = service.get_file(user.tenant_id, user.user_id, file_id)
    except FileNotFoundError:
        return JSONResponse(status_code=404, content={"error": "File not found"})
    return Response(content=content, media_type="text/html")


def _extract_excel(content: bytes) -> list[dict]:
    """Extract Excel sheets as structured data."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            str_row = [str(cell) if cell is not None else "" for cell in row]
            if i == 0:
                headers = str_row
            else:
                rows.append(str_row)
            if i >= 200:  # Limit to 200 rows
                break
        sheets.append({"name": sheet_name, "headers": headers, "rows": rows[:200]})
    wb.close()
    return sheets


@router.delete("/{file_id}")
async def delete_file(file_id: str, user: AuthUser = Depends(get_current_user)):
    """删除文件。"""
    service = get_file_service()

    deleted = service.delete_file(user.tenant_id, user.user_id, file_id)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"File not found: {file_id}")
    return {"status": "ok", "file_id": file_id}


# ── Workspace File API ──

workspace_router = APIRouter(prefix="/api/workspace", tags=["workspace"])


def _get_workspace_dir(tenant_id: str, user_id: str, session_id: str) -> Path:
    """Get workspace directory for a session, validate it exists."""
    from dependencies import get_sandbox_manager
    sandbox = get_sandbox_manager()
    workspace = sandbox.get_workspace(tenant_id, user_id, session_id)
    return Path(workspace)


def _safe_resolve(workspace: Path, rel_path: str) -> Path:
    """Resolve a relative path within workspace, preventing path traversal."""
    resolved = (workspace / rel_path).resolve()
    if not str(resolved).startswith(str(workspace.resolve())):
        raise HTTPException(status_code=403, detail="Path traversal not allowed")
    return resolved


@workspace_router.get("/{session_id}/files")
async def list_workspace_files(
    session_id: str,
    user: AuthUser = Depends(get_current_user),
):
    """列出会话 workspace 中的文件。"""
    workspace = _get_workspace_dir(user.tenant_id, user.user_id, session_id)

    if not workspace.exists():
        return {"session_id": session_id, "files": []}

    files = []
    for f in sorted(workspace.rglob("*")):
        if f.is_file():
            rel = f.relative_to(workspace)
            stat = f.stat()
            files.append({
                "path": str(rel),
                "filename": f.name,
                "size_bytes": stat.st_size,
                "modified_at": stat.st_mtime,
            })

    return {"session_id": session_id, "files": files}


@workspace_router.get("/{session_id}/files/{path:path}/download")
async def download_workspace_file(
    session_id: str,
    path: str,
    user: AuthUser = Depends(get_current_user),
):
    """下载 workspace 文件。"""
    workspace = _get_workspace_dir(user.tenant_id, user.user_id, session_id)
    file_path = _safe_resolve(workspace, path)

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    content = file_path.read_bytes()
    import mimetypes
    content_type = mimetypes.guess_type(file_path.name)[0] or "application/octet-stream"
    encoded_name = quote(file_path.name)

    return Response(
        content=content,
        media_type=content_type,
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
            "Content-Length": str(len(content)),
        },
    )


@workspace_router.get("/{session_id}/files/{path:path}/preview")
async def preview_workspace_file(
    session_id: str,
    path: str,
    user: AuthUser = Depends(get_current_user),
):
    """返回 workspace 文件的结构化预览数据。"""
    workspace = _get_workspace_dir(user.tenant_id, user.user_id, session_id)
    file_path = _safe_resolve(workspace, path)

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail=f"File not found: {path}")

    content = file_path.read_bytes()
    ext = file_path.suffix.lower()
    filename = file_path.name
    base_url = f"/api/workspace/{session_id}/files/{quote(path, safe='/')}"

    # Image
    if ext in ('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp', '.svg'):
        return {"type": "image", "url": f"{base_url}/download", "filename": filename}

    # PDF
    if ext == '.pdf':
        return {"type": "pdf", "url": f"{base_url}/download", "filename": filename}

    # DOCX
    if ext in ('.docx', '.doc'):
        try:
            import io
            from docx import Document as DocxDocument
            doc = DocxDocument(io.BytesIO(content))
            text = "\n".join(p.text for p in doc.paragraphs)
        except Exception:
            text = "(DOCX 解析失败)"
        return {"type": "docx", "content": text, "filename": filename}

    # Excel
    if ext in ('.xlsx', '.xls'):
        try:
            sheets = _extract_excel(content)
            return {"type": "excel", "sheets": sheets, "filename": filename}
        except Exception as e:
            return {"type": "text", "content": f"[Excel 解析失败: {e}]", "filename": filename}

    # HTML
    if ext == '.html':
        try:
            source = content.decode('utf-8')
        except UnicodeDecodeError:
            source = content.decode('utf-8', errors='replace')
        return {"type": "html", "source": source, "filename": filename}

    # Code files
    code_exts = {
        '.py': 'python', '.js': 'javascript', '.ts': 'typescript',
        '.jsx': 'jsx', '.tsx': 'tsx', '.css': 'css', '.scss': 'scss',
        '.java': 'java', '.go': 'go', '.rs': 'rust', '.c': 'c',
        '.cpp': 'cpp', '.h': 'c', '.rb': 'ruby', '.php': 'php',
        '.sh': 'bash', '.sql': 'sql', '.yaml': 'yaml', '.yml': 'yaml',
        '.xml': 'xml',
    }
    if ext in code_exts:
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            text = content.decode('utf-8', errors='replace')
        return {"type": "code", "content": text, "language": code_exts[ext], "filename": filename}

    # Markdown
    if ext == '.md':
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            text = content.decode('utf-8', errors='replace')
        return {"type": "markdown", "content": text, "filename": filename}

    # Plain text
    if ext in ('.txt', '.csv', '.json', '.log', '.ini', '.conf'):
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                text = content.decode('gbk')
            except UnicodeDecodeError:
                text = content.decode('utf-8', errors='replace')
        lang = 'json' if ext == '.json' else 'csv' if ext == '.csv' else ''
        return {"type": "code" if lang else "text", "content": text, "language": lang, "filename": filename}

    # Fallback
    return {
        "type": "unsupported",
        "filename": filename,
        "size_bytes": len(content),
    }
